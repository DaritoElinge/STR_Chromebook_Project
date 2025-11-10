# ======================================================
# VISTAS DE REPORTES (ADMIN)
# ======================================================

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Q
from datetime import datetime
import calendar

# Importar Modelos
from core.models import Usuario
from Gestion_Equipos.models import Reserva, Equipo, AsignacionEquipo

# Importar openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    Workbook = None
    print("ADVERTENCIA: 'openpyxl' no está instalado. La descarga de Excel fallará.")
    print("Ejecuta: pip install openpyxl")


def ver_reportes(request):
    """Vista para visualizar y generar reportes"""
    
    if not request.session.get('usuario_id'):
        messages.error(request, 'Debe iniciar sesión.')
        return redirect('login')
    if request.session.get('usuario_tipo') != 'administrador':
        messages.error(request, 'Acceso denegado.')
        return redirect('dashboard')
    
    usuario = Usuario.objects.get(id_usuario=request.session.get('usuario_id'))
    
    mes_filtro = request.GET.get('mes', datetime.now().month)
    anio_filtro = request.GET.get('anio', datetime.now().year)
    
    try:
        mes_filtro = int(mes_filtro)
        anio_filtro = int(anio_filtro)
    except:
        mes_filtro = datetime.now().month
        anio_filtro = datetime.now().year
    
    # 1. Obtener reservas del mes
    reservas_mes = Reserva.objects.filter(
        fecha_uso__month=mes_filtro,
        fecha_uso__year=anio_filtro
    ).select_related('id_usuario', 'id_carrera', 'id_asignatura', 'id_aula', 'id_aula__id_bloque')
    
    # 2. Estadísticas generales
    total_reservas = reservas_mes.count()
    reservas_aprobadas = reservas_mes.filter(estado_reserva='Aprobada').count()
    reservas_rechazadas = reservas_mes.filter(estado_reserva='Rechazada').count()
    reservas_pendientes = reservas_mes.filter(estado_reserva='Pendiente').count()
    reservas_finalizadas = reservas_mes.filter(estado_reserva__iexact='Finalizada').count()
    total_equipos_solicitados = sum([r.cant_solicitada for r in reservas_mes.filter(estado_reserva='Aprobada')])

    # 3. Reservas por Carrera
    reservas_por_carrera = reservas_mes.values(
        'id_carrera__nom_carrera'
    ).annotate(
        cantidad=Count('id_reserva')
    ).order_by('-cantidad')
    
    # 4. Reservas por Docente
    reservas_por_docente = reservas_mes.values(
        'id_usuario__nom_completo'
    ).annotate(
        cantidad=Count('id_reserva')
    ).order_by('-cantidad')[:10]
    
    # 5. Racks más usados (basado en equipos asignados en reservas FINALIZADAS)
    reservas_finalizadas_ids = reservas_mes.filter(estado_reserva__iexact='Finalizada').values_list('id_reserva', flat=True)
        
    racks_mas_usados = AsignacionEquipo.objects.filter(
        id_reserva_id__in=reservas_finalizadas_ids,
        id_equipo__id_rack__isnull=False # Solo contar equipos que tienen un rack
    ).values(
        'id_equipo__id_rack__nom_rack', # Agrupar por nombre de rack
        'id_equipo__id_rack__ubicacion'
    ).annotate(
        total_equipos=Count('id_equipo') # Contar cuántos equipos de ese rack se usaron
    ).order_by('-total_equipos')

    
    # 6. Selectores de Mes/Año
    meses = [{'num': i, 'nombre': calendar.month_name[i]} for i in range(1, 13)]
    anios = list(range(2024, datetime.now().year + 2))
    
    context = {
        'usuario': usuario, 'mes_filtro': mes_filtro, 'anio_filtro': anio_filtro,
        'mes_nombre': calendar.month_name[mes_filtro], 'meses': meses, 'anios': anios,
        'total_reservas': total_reservas, 'reservas_aprobadas': reservas_aprobadas,
        'reservas_rechazadas': reservas_rechazadas, 'reservas_pendientes': reservas_pendientes, 
        'reservas_finalizadas': reservas_finalizadas,
        'total_equipos_solicitados': total_equipos_solicitados,
        'reservas_por_carrera': reservas_por_carrera,
        'reservas_por_docente': reservas_por_docente,
        'racks_mas_usados': racks_mas_usados, # <-- Añadido al contexto
        'reservas_mes': reservas_mes.order_by('-fecha_uso'),
    }
    
    return render(request, 'administrador/ver_reportes.html', context)


def descargar_reporte_excel(request):
    """Vista para descargar reporte mensual en Excel"""
    
    if request.session.get('usuario_tipo') != 'administrador':
        messages.error(request, 'Acceso denegado.')
        return redirect('dashboard')
    
    if Workbook is None:
        messages.error(request, 'La librería openpyxl no está instalada. No se puede generar el reporte.')
        return redirect('ver_reportes')

    mes = int(request.GET.get('mes', datetime.now().month))
    anio = int(request.GET.get('anio', datetime.now().year))
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {calendar.month_name[mes]} {anio}"
    
    # Estilos
    header_fill = PatternFill(start_color="016BB8", end_color="016BB8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Título
    ws['A1'] = f'REPORTE DE RESERVAS DE CHROMEBOOKS - {calendar.month_name[mes].upper()} {anio}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:L1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws.merge_cells('A2:L2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Obtener TODAS las reservas del mes
    reservas = Reserva.objects.filter(
        fecha_uso__month=mes,
        fecha_uso__year=anio
    ).select_related(
        'id_usuario', 'id_carrera', 'id_asignatura', 'id_aula', 'id_aula__id_bloque'
    ).order_by('fecha_uso', 'hora_inicio')
    
    # --- Estadísticas Generales ---
    current_row = 4
    ws[f'A{current_row}'] = 'ESTADÍSTICAS GENERALES'
    ws[f'A{current_row}'].font = subtitle_font
    current_row += 1
    
    ws[f'A{current_row}'] = 'Total de Reservas:'
    ws[f'B{current_row}'] = reservas.count()
    current_row += 1
    ws[f'A{current_row}'] = 'Aprobadas:'
    ws[f'B{current_row}'] = reservas.filter(estado_reserva='Aprobada').count()
    current_row += 1
    ws[f'A{current_row}'] = 'Rechazadas:'
    ws[f'B{current_row}'] = reservas.filter(estado_reserva='Rechazada').count()
    current_row += 1
    ws[f'A{current_row}'] = 'Pendientes:'
    ws[f'B{current_row}'] = reservas.filter(estado_reserva='Pendiente').count()
    current_row += 1
    ws[f'A{current_row}'] = 'Finalizadas:'
    ws[f'B{current_row}'] = reservas.filter(estado_reserva__iexact='Finalizada').count()
    current_row += 1
    ws[f'A{current_row}'] = 'Total Equipos Solicitados:'
    ws[f'B{current_row}'] = sum([r.cant_solicitada for r in reservas.filter(Q(estado_reserva='Aprobada') | Q(estado_reserva__iexact='Finalizada'))])
    current_row += 1
    
    # --- Racks más usados del mes ---
    current_row += 1  # Espacio en blanco
    reservas_finalizadas_ids = reservas.filter(estado_reserva__iexact='Finalizada').values_list('id_reserva', flat=True)
    racks_mas_usados = AsignacionEquipo.objects.filter(
        id_reserva_id__in=reservas_finalizadas_ids,
        id_equipo__id_rack__isnull=False
    ).values(
        'id_equipo__id_rack__nom_rack'
    ).annotate(
        cantidad=Count('id_equipo')
    ).order_by('-cantidad')
    
    ws[f'A{current_row}'] = 'Racks más usados del mes:'
    ws[f'A{current_row}'].font = subtitle_font
    current_row += 1
    
    if racks_mas_usados:
        for rack in racks_mas_usados:
            ws[f'A{current_row}'] = f"  • {rack['id_equipo__id_rack__nom_rack']}"
            ws[f'B{current_row}'] = f"{rack['cantidad']} equipos"
            current_row += 1
    else:
        ws[f'A{current_row}'] = '  • N/A'
        ws[f'B{current_row}'] = 'Sin datos'
        current_row += 1
    
    # --- Detalle de reservas ---
    current_row += 3
    ws[f'A{current_row}'] = 'DETALLE DE RESERVAS'
    ws[f'A{current_row}'].font = subtitle_font
    
    current_row += 1
    headers = ['Fecha', 'Hora Inicio', 'Hora Fin', 'Docente', 'Carrera', 'Asignatura', 
               'Bloque', 'Aula', 'Cantidad', 'Responsable', 'Teléfono', 'Estado']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header; cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = border
    
    current_row += 1
    for reserva in reservas:
        ws.cell(row=current_row, column=1, value=reserva.fecha_uso.strftime('%d/%m/%Y'))
        ws.cell(row=current_row, column=2, value=reserva.hora_inicio.strftime('%H:%M'))
        ws.cell(row=current_row, column=3, value=reserva.hora_fin.strftime('%H:%M'))
        ws.cell(row=current_row, column=4, value=reserva.id_usuario.nom_completo)
        ws.cell(row=current_row, column=5, value=reserva.id_carrera.nom_carrera)
        ws.cell(row=current_row, column=6, value=reserva.id_asignatura.nom_asignatura)
        ws.cell(row=current_row, column=7, value=reserva.id_aula.id_bloque.nom_bloque)
        ws.cell(row=current_row, column=8, value=reserva.id_aula.nom_aula)
        ws.cell(row=current_row, column=9, value=reserva.cant_solicitada)
        ws.cell(row=current_row, column=10, value=reserva.responsable_entrega)
        ws.cell(row=current_row, column=11, value=reserva.telefono_contacto)
        ws.cell(row=current_row, column=12, value=reserva.estado_reserva)
        
        for col in range(1, 13):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border; cell.alignment = Alignment(horizontal='left', vertical='center')
        
        current_row += 1
    
    # Ajustar ancho de columnas
    column_widths = [12, 12, 12, 30, 35, 30, 10, 15, 10, 35, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Reporte_Chromebooks_{calendar.month_name[mes]}_{anio}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response